import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from docx import Document
import pdfplumber

def read_pdf(file_path):
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text()
    return text

def read_xlsx(file_path):
    wb = load_workbook(filename=file_path)
    sheets = wb.sheetnames
    data = {}
    for sheet in sheets:
        ws = wb[sheet]
        data[sheet] = pd.DataFrame(ws.values)
    return data

def read_csv(file_path):
    return pd.read_csv(file_path)

def read_docx(file_path):
    doc = Document(file_path)
    text = "\n".join([para.text for para in doc.paragraphs])
    return text

def read_file(file_path):
    if file_path.endswith('.pdf'):
        return read_pdf(file_path)
    elif file_path.endswith('.xlsx'):
        return read_xlsx(file_path)
    elif file_path.endswith('.csv'):
        return read_csv(file_path)
    elif file_path.endswith('.docx'):
        return read_docx(file_path)
    else:
        raise ValueError("Unsupported file format")
